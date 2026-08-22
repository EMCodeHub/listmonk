#!/usr/bin/env python3
"""Build a strict, country-segmented air-conditioning audience locally."""

from __future__ import annotations

import csv
import json
import math
import re
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path

import dns.exception
import dns.resolver
from email_validator import EmailNotValidError, validate_email
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "clear_emails_air_condition" / "All_Air_Condition_Leads.xlsx"
OUTPUT = ROOT / "clean_emails_air_condition"
BOUNCES = ROOT / "listmonk_emails" / ".air_condition_vps_bounces.csv"
DOMAIN_EVIDENCE = ROOT / "listmonk_emails" / ".strict_vps_domains.csv"
MANUAL_SUPPRESSIONS = ROOT / "listmonk_emails" / "manual_suppressions.json"
CHECKPOINT = ROOT / "listmonk_emails" / ".air_condition_dns_checkpoint.csv"

EMAIL_KEYS = ("email", "emails", "email_address", "e_mail", "mail")
COUNTRY_KEYS = ("country", "country_name", "pais", "país")
NAME_KEYS = ("name", "business_name", "company", "company_name", "title")
PHONE_KEYS = ("telephone", "phone", "phone_number", "tel")
WEBSITE_KEYS = ("website_instagram_facebook", "website", "url", "site")

EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+$")
ASSET_RE = re.compile(r"\.(?:png|jpe?g|gif|svg|webp|css|js|ico|woff2?|ttf|pdf|xml|php|html?)$", re.I)
PLACEHOLDER_RE = re.compile(
    r"(?:^|[._-])(?:test|testing|teste|prueba|example|exemple|esempio|exemplo|sample|dummy|fake|falso|invalid|asdf|qwerty|unknown|null|none|spam|demo|domain|mailbox|john[._-]?doe|jane[._-]?doe)(?:$|[._-])",
    re.I,
)
HEX_RE = re.compile(r"^[0-9a-f]{20,}$", re.I)
UUID_RE = re.compile(r"^[0-9a-f]{8}-?[0-9a-f]{4}-?[0-9a-f]{4}-?[0-9a-f]{4}-?[0-9a-f]{12}$", re.I)
ENCODED_RE = re.compile(r"(?:u00[0-9a-f]{2}|%[0-9a-f]{2}|x[0-9a-f]{2})", re.I)
REPEATED_RE = re.compile(r"(.)\1{4,}", re.I)
DISPOSABLE = {"10minutemail.com", "guerrillamail.com", "mailinator.com", "tempmail.com", "temp-mail.org", "yopmail.com", "throwawaymail.com", "trashmail.com", "maildrop.cc", "sharklasers.com", "getnada.com", "dispostable.com"}
BAD_DOMAINS = {"example.com", "example.org", "example.net", "localhost", "email.com", "correo.com", "mail.com.br", "yourmail.com", "website.com", "mysite.com"}
ROLE_LOCALS = {
    "abuse", "admin", "billing", "booking", "careers", "commercial", "comercial", "contact", "contacto", "contato",
    "customer", "customerservice", "email", "enquiries", "finance", "geral", "hello", "help", "hr", "info", "jobs",
    "mail", "marketing", "office", "orders", "postmaster", "privacy", "reception", "reservas", "root", "sales", "service",
    "servicio", "shop", "soporte", "support", "suporte", "webmaster", "noreply", "no-reply", "donotreply", "do-not-reply",
    "mailer-daemon", "newsletter", "unsubscribe",
}
COMPACT_ROLES = {re.sub(r"[^a-z0-9]", "", value) for value in ROLE_LOCALS}
FINAL_DNS = {"valid_mx", "dns_nxdomain", "dns_null_mx", "dns_no_mx"}


def key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (value or "").strip().lower()).strip("_")


def normalize(raw: str) -> str:
    value = (raw or "").strip().lower().replace("\ufeff", "")
    while value.startswith("%20"):
        value = value[3:]
    return value.removeprefix("mailto:").strip(" \t\r\n,;:'\"<>[]()")


def entropy(value: str) -> float:
    counts = Counter(value)
    return -sum((n / len(value)) * math.log2(n / len(value)) for n in counts.values()) if value else 0.0


def field(row: dict[str, str], aliases: tuple[str, ...]) -> str:
    normalized = {key(name): value for name, value in row.items() if name is not None}
    return next((normalized[name] for name in aliases if normalized.get(name)), "")


def load_evidence() -> tuple[set[str], set[str]]:
    known: set[str] = set()
    with BOUNCES.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle, delimiter="|"):
            email = normalize(row.get("email", ""))
            if email:
                known.add(email)
    if MANUAL_SUPPRESSIONS.exists():
        payload = json.loads(MANUAL_SUPPRESSIONS.read_text(encoding="utf-8"))
        known.update(normalize(email) for email in payload.get("emails", []) if normalize(email))
    risky: set[str] = set()
    with DOMAIN_EVIDENCE.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle, delimiter="|"):
            attempts, bounces = int(row["attempts"]), int(row["bounces"])
            rate = bounces / attempts if attempts else 0
            if (attempts >= 3 and rate >= .20) or (attempts >= 10 and rate >= .15) or (attempts >= 25 and rate >= .10):
                risky.add(row["domain"].strip().lower())
    return known, risky


def reject_reason(email: str, known: set[str], risky: set[str]) -> str | None:
    if email in known:
        return "known_vps_bounce_or_suppression"
    if email.count("@") != 1 or not EMAIL_RE.fullmatch(email):
        return "invalid_syntax"
    try:
        validate_email(email, check_deliverability=False, allow_smtputf8=True)
    except EmailNotValidError:
        return "invalid_syntax"
    local, domain = email.rsplit("@", 1)
    compact = re.sub(r"[^a-z0-9]", "", local)
    if domain in BAD_DOMAINS or domain in DISPOSABLE or PLACEHOLDER_RE.search(domain):
        return "placeholder_or_disposable_domain"
    if domain in risky:
        return "observed_high_bounce_domain"
    if local in ROLE_LOCALS or compact in COMPACT_ROLES:
        return "generic_or_technical_role"
    if PLACEHOLDER_RE.search(local):
        return "placeholder_local"
    if len(local) <= 4 or len(local) >= 40:
        return "implausible_local_length"
    if any(ch.isdigit() for ch in local):
        return "numeric_or_id_local"
    if any(ch in local for ch in "-_+"):
        return "risky_separator_or_alias"
    if ASSET_RE.search(email):
        return "asset_false_positive"
    if HEX_RE.fullmatch(compact) or UUID_RE.fullmatch(local) or ENCODED_RE.search(local):
        return "encoded_or_machine_token"
    if REPEATED_RE.search(compact):
        return "implausible_repetition"
    if len(compact) >= 20 and compact == local and entropy(compact) >= 3.6:
        return "random_high_entropy_token"
    return None


def load_cache() -> dict[str, tuple[str, str]]:
    cache: dict[str, tuple[str, str]] = {}
    for path in (ROOT / "listmonk_emails").glob("clean_emails_*/domain_validation_cache.csv"):
        with path.open(encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                domain = row.get("domain", "").strip().lower().rstrip(".")
                if domain:
                    cache[domain] = (row.get("status", ""), row.get("detail", ""))
    if CHECKPOINT.exists():
        with CHECKPOINT.open(encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                cache[row["domain"].strip().lower()] = (row["status"], row.get("detail", ""))
    return cache


def check_mx(domain: str) -> tuple[str, str]:
    resolver = dns.resolver.Resolver(configure=True)
    resolver.timeout, resolver.lifetime = 3.0, 5.0
    try:
        answers = resolver.resolve(domain, "MX")
        if any(int(record.preference) == 0 and str(record.exchange) == "." for record in answers):
            return "dns_null_mx", "."
        return "valid_mx", ",".join(sorted(str(record.exchange).rstrip(".") for record in answers))
    except dns.resolver.NXDOMAIN:
        return "dns_nxdomain", "NXDOMAIN"
    except dns.resolver.NoAnswer:
        return "dns_no_mx", "NoAnswer"
    except (dns.resolver.NoNameservers, dns.exception.Timeout) as exc:
        return "dns_inconclusive", type(exc).__name__


def write_checkpoint(cache: dict[str, tuple[str, str]], domains: set[str]) -> None:
    with CHECKPOINT.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle); writer.writerow(["domain", "status", "detail"])
        writer.writerows((domain, *cache[domain]) for domain in sorted(domains) if domain in cache)


def source_rows() -> tuple[list[str], object]:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    for row_number, values in enumerate(rows, 1):
        headers = [str(value).strip() if value is not None else "" for value in values]
        normalized = {key(value) for value in headers}
        if normalized.intersection(EMAIL_KEYS) and normalized.intersection(COUNTRY_KEYS):
            def records():
                try:
                    for values in rows:
                        yield {headers[index]: "" if value is None else str(value) for index, value in enumerate(values) if index < len(headers)}
                finally:
                    workbook.close()
            return headers, records()
    workbook.close()
    raise SystemExit("Workbook must contain recognizable email and country headers")


def main() -> None:
    started = time.time()
    if OUTPUT.exists():
        raise SystemExit(f"Refusing to overwrite existing output: {OUTPUT}")
    for required in (SOURCE, BOUNCES, DOMAIN_EVIDENCE):
        if not required.exists():
            raise SystemExit(f"Missing required input: {required}")
    known, risky = load_evidence()
    candidates: dict[str, dict[str, dict[str, str]]] = {}
    input_by_country, reasons = Counter(), Counter()
    removed: list[tuple[str, str, str]] = []
    domains: set[str] = set()
    headers, reader = source_rows()
    normalized_headers = {key(name) for name in headers}
    has_email_column = bool(normalized_headers.intersection(EMAIL_KEYS))
    for row in reader:
            country = field(row, COUNTRY_KEYS).strip()
            # Some exports have no dedicated EMAIL column. In that case only
            # extract explicit address-shaped tokens already present in fields;
            # URLs are never guessed or crawled during a cleanup operation.
            raw_email = field(row, EMAIL_KEYS) if has_email_column else " ".join(
                value for value in row.values() if "@" in (value or "")
            )
            if not country:
                for token in re.split(r"[;,\s]+", raw_email or ""):
                    if normalize(token):
                        reasons["missing_country"] += 1
                continue
            bucket = candidates.setdefault(country, {})
            for token in re.split(r"[;,\s]+", raw_email or ""):
                email = normalize(token)
                if not email:
                    continue
                input_by_country[country] += 1
                reason = reject_reason(email, known, risky)
                if reason:
                    reasons[reason] += 1; removed.append((country, email, reason)); continue
                if email in bucket:
                    reasons["duplicate_within_country"] += 1; removed.append((country, email, "duplicate_within_country")); continue
                bucket[email] = row; domains.add(email.rsplit("@", 1)[1])

    cache = load_cache()
    to_check = sorted(domain for domain in domains if cache.get(domain, ("", ""))[0] not in FINAL_DNS)
    # Keep concurrency deliberately moderate. Large bursts can saturate the
    # local resolver and incorrectly turn healthy domains into inconclusive
    # results; inconclusive domains are excluded under the strict policy.
    with ThreadPoolExecutor(max_workers=20) as pool:
        futures = {pool.submit(check_mx, domain): domain for domain in to_check}
        for index, future in enumerate(as_completed(futures), 1):
            domain = futures[future]
            try:
                cache[domain] = future.result()
            except Exception as exc:
                cache[domain] = ("dns_inconclusive", type(exc).__name__)
            if index % 250 == 0:
                write_checkpoint(cache, domains)
    write_checkpoint(cache, domains)

    global_seen: set[str] = set()
    kept: dict[str, list[tuple[str, dict[str, str]]]] = {}
    for country in sorted(candidates):
        kept[country] = []
        for email, row in sorted(candidates[country].items()):
            status = cache.get(email.rsplit("@", 1)[1], ("dns_unknown", ""))[0]
            if status != "valid_mx":
                reasons[status or "dns_unknown"] += 1; removed.append((country, email, status or "dns_unknown")); continue
            if email in global_seen:
                reasons["duplicate_across_countries"] += 1; removed.append((country, email, "duplicate_across_countries")); continue
            global_seen.add(email); kept[country].append((email, row))

    OUTPUT.mkdir(parents=True)
    summary_rows: list[dict[str, object]] = []
    for rank, country in enumerate(sorted(kept, key=lambda name: (-len(kept[name]), name.casefold())), 1):
        records = kept[country]
        summary_rows.append({"rank": rank, "country": country, "input_nonempty": input_by_country[country], "kept": len(records), "removed": input_by_country[country] - len(records)})
        if not records:
            continue
        safe_country = re.sub(r'[<>:"/\\|?*]+', "_", country).strip(" .") or "Unknown"
        folder = OUTPUT / f"{rank:03d} - {safe_country} ({len(records)})"; folder.mkdir()
        (folder / f"emails({len(records)}).txt").write_text("\n".join(email for email, _ in records) + ("\n" if records else ""), encoding="utf-8")
        with (folder / "listmonk_subscribers.csv").open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=["email", "name", "attributes"], quoting=csv.QUOTE_ALL); writer.writeheader()
            for email, row in records:
                attributes = {"telephone": field(row, PHONE_KEYS), "website": field(row, WEBSITE_KEYS), "country": country, "sector": "air_conditioning"}
                writer.writerow({"email": email, "name": field(row, NAME_KEYS), "attributes": json.dumps(attributes, ensure_ascii=False, separators=(",", ":"))})
    with (OUTPUT / "country_summary.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["rank", "country", "input_nonempty", "kept", "removed"]); writer.writeheader(); writer.writerows(summary_rows)
    with (OUTPUT / "domain_validation_cache.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle); writer.writerow(["domain", "status", "detail"]); writer.writerows((d, *cache.get(d, ("dns_unknown", ""))) for d in sorted(domains))
    with (OUTPUT / "removed_email_audit.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle); writer.writerow(["country", "email", "reason"]); writer.writerows(removed)
    total_input, total_kept = sum(input_by_country.values()), len(global_seen)
    summary = {"generated_utc": datetime.now(timezone.utc).isoformat(), "source": SOURCE.name, "countries": len(kept), "input_nonempty_email_entries": total_input, "kept_unique_emails": total_kept, "removed_entries": total_input-total_kept, "known_vps_bounces_loaded": len(known), "risky_domains_from_vps": len(risky), "candidate_domains": len(domains), "dns_live_checks": len(to_check), "reason_counts": dict(reasons.most_common()), "processing_seconds": round(time.time()-started, 1)}
    (OUTPUT / "validation_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2)+"\n", encoding="utf-8")
    report = ["# Air-conditioning email cleanup", "", "Strict country-level filtering using production bounce history and confirmed MX routing.", "", f"- Input entries: **{total_input:,}**", f"- Kept unique emails: **{total_kept:,}**", f"- Removed: **{total_input-total_kept:,}**", f"- Countries: **{len(kept):,}**", f"- Known production bounces loaded: **{len(known):,}**", "", "## Removal signals", "", "| Signal | Matches |", "|---|---:|"]
    report += [f"| {reason} | {count:,} |" for reason, count in reasons.most_common()]
    report += ["", "Only confirmed MX domains were retained. SMTP mailbox verification was deliberately not attempted; an MX record cannot guarantee that an individual mailbox exists.", ""]
    (OUTPUT / "VALIDATION_REPORT.md").write_text("\n".join(report), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
